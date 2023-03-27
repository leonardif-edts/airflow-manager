import json
import logging
from datetime import datetime
from typing import Union, Optional

import openpyxl
import gspread

from scripts import const, utils
from scripts.modules import project


# Public
def extract_tf(version_dir: str, filename: Optional[str] = None, sheet_id: Optional[str] = None):
    if (filename and sheet_id is None) or (filename is None and sheet_id):
        mode = "xlsx" if (filename) else "sheet"
        id = filename if (filename) else sheet_id
    else:
        raise AssertionError

    # Authenticate SA
    if (mode == "sheet"):
        logging.info(f"Authenticate SA: `{json.load(open('service-account.json')).get('client_email')}`")
        client = gspread.service_account("service-account.json")
    
    # Get workbook
    wb = openpyxl.load_workbook(filename, data_only=True) if (mode == "xlsx") else client.open_by_key(sheet_id)
    config = {}

    logging.info(f"Extract project deployment configuration from {f'file: {filename}' if (mode == 'xlsx') else f'sheet_id: {sheet_id}'}")
    config = _extract_metadata(wb, mode, config)

    ws_index = wb["INDEX"] if (mode == "xlsx") else wb.worksheet("INDEX")
    data = ws_index.iter_rows(min_col=1, min_row=3, values_only=True) if (mode == "xlsx") else utils.slicing_list(ws_index.get_all_values(), min_col=1, min_row=3)
    for row in data:
        bq_tablename = row[1]
        dag_id = row[5]
        if (bq_tablename is None):
            continue
        
        logging.info(f"Extract DAG deployment configuration: Table `{bq_tablename}` DAG `{dag_id}`")
        config = _extract_dag_config(wb, mode, config, row)

    timestamp = datetime.now()
    filename = f"{timestamp.strftime('%Y%m%d%H%M%S')}.json"
    utils.export_json(config, version_dir, filename)
    project.update_plan_logs(config, mode, id, filename)
    project.update_metadata({
        "plan.latest_version": filename.replace(".json", ""),
        "plan.update_ts": timestamp
    })


# Private
def _extract_metadata(wb: Union[openpyxl.Workbook, gspread.spreadsheet.Spreadsheet], mode: str, config: dict) -> dict:
    ws = wb["METADATA"] if (mode == "xlsx") else wb.worksheet("METADATA")
    config["project"] = config.get("project", {"dev":{}, "prd":{}})
    config["project"]["dev"] = config["project"].get("dev", {})
    config["project"]["prd"] = config["project"].get("prd", {})

    data = ws.iter_rows(min_col=1, min_row=3, values_only=True) if (mode == "xlsx") else utils.slicing_list(ws.get_all_values(), min_col=1, min_row=3)
    for row in data:
        p_key, p_type, p_dev, p_prd = row

        p_key_mapped = p_key.strip(" ").replace(" ", "_").lower()
        config["project"]["dev"][p_key_mapped] = _extract_params(p_type, p_dev)
        config["project"]["prd"][p_key_mapped] = _extract_params(p_type, p_prd)

    return config

def _extract_params(datatype: str, value: str) -> Union[str, dict, list]:
    if (datatype == "string"):
        return value.strip()

def _extract_dag_config(wb: Union[openpyxl.Workbook, gspread.spreadsheet.Spreadsheet], mode: str, config: dict, row: list) -> dict:
    dag_config = _extract_dag_index(row)
    
    tablename = dag_config["bq_tablename"]
    dag_config = _extract_dag_table(wb, mode, dag_config, tablename)

    dags = config.get("dags", [])
    dags.append(dag_config)
    config["dags"] = dags
    return config

def _extract_dag_index(row: list) -> dict:
    dag_config = {
        key: cell if (key not in const.INDEX_COLUMNS_ARRAY) else cell.split(";")
        for key, cell in zip(const.INDEX_COLUMNS, row)
    }
    return dag_config

def _extract_dag_table(wb: Union[openpyxl.Workbook, gspread.spreadsheet.Spreadsheet], mode: str, dag_config: dict, tablename: str) -> dict:
    table_columns = dag_config.get("table_columns", {})
    ext_cols = _extract_dag_table_value(wb, mode, tablename, ["name", "datatype"], min_row=3, min_col=2, max_col=3)
    src_cols = _extract_dag_table_value(wb, mode, tablename, ["name", "datatype", "transformation"], min_row=3, min_col=6, max_col=8)
    stg_cols = _extract_dag_table_value(wb, mode, tablename, ["name", "datatype", "source", "transformation"], min_row=3, min_col=10, max_col=13)
    dw_cols  = _extract_dag_table_value(wb, mode, tablename, ["name", "datatype", "unique", "partition", "cluster"], min_row=3, min_col=15, max_col=19)

    table_columns = {
        "ext": table_columns.get("ext", ext_cols),
        "src": table_columns.get("src", src_cols),
        "stg": table_columns.get("stg", stg_cols),
        "dw": table_columns.get("dw", dw_cols)
    }

    dag_config["unique"] = utils.extend_coalesce(utils.get_filtered_columns(dw_cols, "unique", [True, "TRUE"]))
    dag_config["partition"] = utils.get_filtered_columns(dw_cols, "partition", [True, "TRUE"])[0]
    dag_config["cluster"] = utils.get_filtered_columns(dw_cols, "cluster", [True, "TRUE"])
    table_columns["dw"] = utils.filter_out_keys(table_columns.get("dw", {}), ["unique", "partition", "cluster"])
    dag_config["columns"] = table_columns
    return dag_config

def _extract_dag_table_value(
        wb: Union[openpyxl.Workbook, gspread.spreadsheet.Spreadsheet],
        mode: str,
        tablename: str,
        column_labels: list,
        min_row: int = 0,
        max_row: int = None,
        min_col: int = 0,
        max_col: int = None
    ):
        if (mode == "xlsx"):
            ws = wb[tablename]
            max_row = utils.coalesce(max_row, ws.max_row)
            max_col = utils.coalesce(max_col, ws.max_column)
        elif (mode == "sheet"):
            ws = wb.worksheet(tablename)

        table_cols = []
        data = ws.iter_rows(min_row, max_row, min_col, max_col, values_only=True) if (mode == "xlsx") else utils.slicing_list(ws.get_all_values(), min_row, max_row, min_col, max_col)
        for row in data:
            if (row[0] is not None and row[0] != "") and (row[0] not in const.EXCLUDE_TABLE_COLUMNS):
                column = {
                    key: col
                    for key, col in zip(column_labels, row)
                }
                table_cols.append(column)
        return table_cols    