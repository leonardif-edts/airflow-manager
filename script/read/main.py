import json
from typing import Union

import openpyxl


def _coalesce(*values):
    return next((v for v in values if v is not None), None)

def _get_filtered_columns(data: list, key: str, value) -> list:
    fltr_columns = [
        column["name"]
        for column in data
        if column[key] == value
    ]
    return fltr_columns

def _filter_out_keys(data: list, keys: list) -> list:
    fltr_columns = [
        {
            key: value
            for key, value in column.items()
            if (key not in keys)
        }
        for column in data
    ]
    return fltr_columns

PARAMS_KEY_MAPPER = {
    "Project ID": "project_id",
    "GCS Bucket": "gcs_bucket"
}

INDEX_COLUMNS = (
    "no",
    "bq_tablename",
    "type",
    "method",
    "source_type",
    "source_database",
    "source_tablename",
    "source_filename",
    "source_connection",
    "dag_type",
    "dag_id"
)

EXCLUDE_COLUMNS = {
    "job_date",
    "load_datetime",
    "job_id",
    "path_filename",
    "row"
}


def extract_metadata(wb: openpyxl.Workbook, config: dict) -> dict:
    ws = wb["METADATA"]
    config["project"] = config.get("project", {"dev":{}, "prd":{}})
    config["project"]["dev"] = config["project"].get("dev", {})
    config["project"]["prd"] = config["project"].get("prd", {})

    for row in ws.iter_rows(min_col=1, min_row=3, values_only=True):
        p_key, p_type, p_dev, p_prd = row

        p_key_mapped = PARAMS_KEY_MAPPER.get(p_key, p_key)
        config["project"]["dev"][p_key_mapped] = extract_params(p_type, p_dev)
        config["project"]["prd"][p_key_mapped] = extract_params(p_type, p_prd)

    return config

def extract_params(datatype: str, value: str) -> Union[str, dict, list]:
    if (datatype == "string"):
        return value.strip()

def extract_dag_config(wb: openpyxl.Workbook, config: dict, row: list) -> dict:
    dag_config = extract_dag_index(row)
    
    tablename = dag_config["bq_tablename"]
    dag_config = extract_dag_table(wb, dag_config, tablename)

    dags = config.get("dags", [])
    dags.append(dag_config)
    config["dags"] = dags
    return config


def extract_dag_index(row: list) -> dict:
    dag_config = {
        key: cell
        for key, cell in zip(INDEX_COLUMNS, row)
    }
    return dag_config


def extract_dag_table(wb: openpyxl.Workbook, dag_config: dict, tablename: str) -> dict:
    table_columns = dag_config.get("table_columns", {})
    ext_cols = extract_dag_table_value(wb, tablename, ["name", "datatype"], min_row=3, min_col=2, max_col=3)
    src_cols = extract_dag_table_value(wb, tablename, ["name", "datatype", "transformation"], min_row=3, min_col=6, max_col=8)
    stg_cols = extract_dag_table_value(wb, tablename, ["name", "datatype", "transformation"], min_row=3, min_col=10, max_col=12)
    dw_cols  = extract_dag_table_value(wb, tablename, ["name", "datatype", "unique", "partition", "cluster"], min_row=3, min_col=14, max_col=18)

    table_columns = {
        "ext": table_columns.get("ext", ext_cols),
        "src": table_columns.get("src", src_cols),
        "stg": table_columns.get("stg", stg_cols),
        "dw": table_columns.get("dw", dw_cols)
    }

    dag_config["unique"] = _get_filtered_columns(dw_cols, "unique", True)
    dag_config["partition"] = _get_filtered_columns(dw_cols, "partition", True)
    dag_config["cluster"] = _get_filtered_columns(dw_cols, "cluster", True)
    table_columns["dw"] = _filter_out_keys(table_columns.get("dw", {}), ["unique", "partition", "cluster"])
    dag_config["colums"] = table_columns
    return dag_config


def extract_dag_table_value(
        wb: openpyxl.Workbook,
        tablename: str,
        column_labels: list,
        min_row: int = 0,
        max_row: int = None,
        min_col: int = 0,
        max_col: int = None
    ):
        ws = wb[tablename]
        max_row = _coalesce(max_row, ws.max_row)
        max_col = _coalesce(max_col, ws.max_column)

        table_cols = []
        for row in ws.iter_rows(min_row, max_row, min_col, max_col, values_only=True):
            if (row[0] is not None) and (row[0] not in EXCLUDE_COLUMNS):
                column = {
                    key: col
                    for key, col in zip(column_labels, row)
                }
                table_cols.append(column)
        return table_cols


if __name__ == "__main__":
    wb = openpyxl.load_workbook("test.xlsx", data_only=True)
    config = {}
    
    print("Extract Project configuration")
    config = extract_metadata(wb, config)
    
    ws_index = wb["INDEX"]
    for i, row in enumerate(ws_index.iter_rows(min_col=1, min_row=3, values_only=True)):
        bq_tablename = row[1]
        dag_id = row[10]

        print(f"Extract DAG configuration: Table `{bq_tablename}` DAG `{dag_id}`")
        config = extract_dag_config(wb, config, row)
    
    with open("config.json", "w") as file:
        file.write(json.dumps(config, indent=2))